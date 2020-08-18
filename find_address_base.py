import openpyxl


# import other_function_xlsx as fxb
# import config_data


class RegionDistrict:
    """
    класс определяющий принадлежность к федеральному округу по строке
    синглтон по сути, но пока реализован как обычный
    """
    # блок входных данных
    country = "Россия"
    support_dir = "support_tables"
    address_file = "address_structure.xlsx"
    address_base_file = support_dir + "\\" + address_file
    district_tuple = ("Центральный", "Приволжский", "Дальневосточный", "Северо-Западный",
                      "Северо-Кавказский", "Сибирский", "Уральский", "Южный", "Не определен")
    region_name_tuple = None  # чуть позже создадим кортеж из списка для ускорения процессов
    post_index_dict = dict()
    region_base = dict()
    region_name_list = list()
    cities_name_dict = dict()
    vehicle_code_dict = dict()

    def __init__(self):

        self.country_name = RegionDistrict.country
        RegionDistrict.make_region_base()

    @classmethod
    def make_region_base(cls):
        # в каких столбиках какие данные
        head_district_number = 5
        head_region_name = 2
        head_post_index = 3
        head_vehicle_code = 4

        # формируем структуру данных для адресов
        # открываем файл со структурой
        district_file = openpyxl.load_workbook(RegionDistrict.address_base_file)
        district_file.active = 0
        # перебираем файл, составляем рабочие списки и словари
        for i in range(2, district_file.active.max_row + 1):
            this_post_index = district_file.active.cell(row=i, column=head_post_index).value
            this_vehicle_code = district_file.active.cell(row=i, column=head_vehicle_code).value
            # print("переменная This_vehicle_code:", this_vehicle_code)
            this_district_number = district_file.active.cell(row=i, column=head_district_number).value
            this_region_name = district_file.active.cell(row=i, column=head_region_name).value
            # заполняем словарь класса с регионами
            cls.region_base[i - 2] = this_region_name, this_district_number, this_vehicle_code, this_post_index
            # FIXME из предыдущей строки нужно убрать что не будем использовать
            #  первые кандидаты this_vehicle_code и this_post_index, так как по этим полям идет поиск региона
            # соберем словарь номеров кодов регионов для ТС
            for one_vehicle_code in this_vehicle_code.split(", "):
                cls.vehicle_code_dict[one_vehicle_code] = this_region_name
            # заодно заполняем словарь почтовых индексов
            cls.post_index_dict[this_post_index] = i - 2

        # переключаемся на вкладку с названиями регионов и собираем список названий
        # (используем для принадлежности городов)
        district_file.active = 1
        for i in range(2, district_file.active.max_row + 1):
            cls.region_name_list.append(district_file.active.cell(row=i, column=1).value)
            # Заодно создадим словарь регионов с пустыми списками городов
            cls.cities_name_dict[i - 2] = set()
        # переключаемся на вкладку с названиями городов и собираем словарь регионов со списками городов
        district_file.active = 2

        for i in range(2, district_file.active.max_row + 1):
            this_city = district_file.active.cell(row=i, column=1).value
            this_region_id = district_file.active.cell(row=i, column=2).value
            # this_district_id = district_file.active.cell(row=i, column=3).value
            cls.cities_name_dict[this_region_id].add(this_city)
        cls.region_name_tuple = tuple(cls.region_name_list)
        district_file.close()

    @classmethod
    def get_region_and_district_on_region_id(cls, region_id):

        region_name = cls.region_base[region_id][0]
        district_number = cls.region_base[region_id][1]
        district_name = cls.district_tuple[district_number]

        return region_name, district_name

    @classmethod
    def find_address(cls, any_address_string):
        region_name = district_name = city_name = None
        flag = 0  # пока регион не найден
        # сначала будем искать по почтовому индексу - найдем индекс в строке
        post_index = cls.search_post_index(any_address_string)
        region_id = cls.post_index_dict.get(post_index)
        if region_id is not None:
            # если почтовый индекс нашли, запоминаем имя региона и федерального округа
            region_name, district_name = cls.get_region_and_district_on_region_id(region_id)
            city_name = cls.search_city_name_in_region(any_address_string, region_id)
            if district_name is not None:
                return cls.country, region_name, district_name, city_name, True

        # если по почтовому индексу ничего нет (или не указан индекс) будем искать по имени области
        else:

            # print(any_address_string)
            region_id = cls.search_region_id(any_address_string)
            if region_id is not None:
                # если почтовый индекс нашли, запоминаем имя региона и федерального округа
                region_name, district_name = cls.get_region_and_district_on_region_id(region_id)
                city_name = cls.search_city_name_in_region(any_address_string, region_id)
                if district_name is not None:
                    return cls.country, region_name, district_name, city_name, True


        # если и по области ничего нет - надо искать по названию города (как повезет)
        region_id, city_name = cls.search_region_on_city_name(any_address_string)
        if region_id is not None:
            # если нашли какой-то город, запоминаем имя региона и федерального округа
            region_name, district_name = cls.get_region_and_district_on_region_id(region_id)
            if district_name is not None:
                return cls.country, region_name, district_name, city_name, False
        else:
                return None, region_name, district_name, city_name, False

    @staticmethod
    def search_post_index(any_address_string):
        for i in range(len(any_address_string) - 6):
            if any_address_string[i: i + 6].isdigit():
                return any_address_string[i: i + 3]

    @classmethod
    def search_region_id(cls, any_address_string):
        for current_id in range(len(cls.region_name_tuple)):
            for region_name in cls.region_name_tuple[current_id].split(" / "):
                if region_name.strip() in any_address_string:
                    region_id = current_id
                    return region_id

    @classmethod
    def search_region_on_city_name(cls, any_address_string):
        for current_region_id, city_names_of_region in cls.cities_name_dict.items():
            for current_city_name in city_names_of_region:
                if current_city_name in any_address_string:
                    region_id = current_region_id
                    city_name = current_city_name
                    # print("найден по городу город:", city_name)
                    return region_id, city_name
        return None, None

    @classmethod
    def search_city_name_in_region(cls, any_address_string: str, current_region_id):
        if current_region_id is not None:
            """ отладка
            print(any_address_string)
            print("Внутренний номер региона:", current_region_id)
            print("Соотвествующие города")
            print(cls.cities_name_dict[current_region_id])
            """
            city_names_of_region = cls.cities_name_dict[current_region_id]
            for current_city_name in city_names_of_region:
                if current_city_name.upper() in any_address_string.upper():
                    # print("найден город:", current_city_name)
                    return current_city_name
        else:
            return None
